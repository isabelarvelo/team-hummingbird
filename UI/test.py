import streamlit as st
import pandas as pd
from PIL import Image
from io import BytesIO
import base64
import openpyxl
from openpyxl.styles import PatternFill
from transformers import pipeline
import numpy as np

def show_overview():
    st.title("Welcome to ReTeach!")

    # Display an image 
    # Ensure the image file is in the same directory as your script, or provide the full path.
    image = Image.open('GPT_LOGO.png')
    col1, col2, col3 = st.columns([1,2,1])

    with col2:
        st.image(image, caption='Enhancing Teaching with Technology', width=300)

    st.markdown("""
        ## Empowering Educators with Data-Driven Insights

        **ReTeach** is a revolutionary platform that listens to classroom interactions 
        and provides real-time, AI-powered feedback to educators. Our goal is to 
        enhance teaching effectiveness and improve learning outcomes.

        ### How ReTeach Works:
        - **Upload Transcripts:** Teachers upload their classroom interaction transcripts.
        - **AI Analysis:** Our AI model analyzes the interactions and suggests improvements.
        - **Feedback & Growth:** Teachers receive personalized feedback for professional development.

        Explore the sidebar to navigate through the application's features and start your journey towards enhanced educational experiences!
    """)



def color_map(val):
    """
    Takes a scalar and returns a string with
    the css property `'background-color'` for a color.
    Uses a non-linear scale for color mapping.
    """
    if pd.isna(val):
        return None
    elif val < 0.2:
        return 'FFFFCC'  # light yellow
    elif val < 0.3:
        return 'D9F0A3'  # light green
    elif val < 0.4:
        return 'ADDD8E'  # green
    elif val < 0.5:
        return '78C679'  # darker green
    else:
        return '31A354'  # dark green
    
# Function to convert DataFrame to Excel and then encode it for download
def apply_excel_color_styles(df, color_map_function, target_columns):
    output = BytesIO()
    writer = pd.ExcelWriter(output, engine='openpyxl')
    df.to_excel(writer, index=False)
    workbook = writer.book
    worksheet = writer.sheets['Sheet1']
    
    for idx, row in df.iterrows():
        for col_name, value in row.items():
            # Apply coloring only to specified columns
            if col_name in target_columns:
                cell = worksheet.cell(row=idx + 2, column=df.columns.get_loc(col_name) + 1)
                color = color_map_function(value)
                if color:
                    cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

    writer.save()
    return output.getvalue()

def get_table_download_link(df, target_columns):
    processed_data = apply_excel_color_styles(df, color_map, target_columns)
    
    b64 = base64.b64encode(processed_data).decode()
    href = f'<a href="data:application/octet-stream;base64,{b64}" download="processed_data.xlsx">Download processed data as Excel</a>'
    return href


# Main function for the Streamlit app
def show_ai_assisted_labeling():
    st.title("Classroom Transcript Labeling Tool")
    st.write("Upload a transcript and get it labeled based on classroom interaction categories.")

    uploaded_file = st.file_uploader("Choose a file")

    if uploaded_file is not None:
        df = pd.read_csv(uploaded_file)
        if 'Legal' in df.columns:
            df.drop('Legal', axis=1, inplace=True)
        
    
        # Initialize the zero-shot classifier
        classifier = pipeline("zero-shot-classification", model="sileod/deberta-v3-base-tasksource-nli")
        
        # Define candidate labels
        candidate_labels = ['PRS', 'REP', 'OTR', 'NEU']

        # Initialize columns for scores
        score_columns = ['PRS_Score', 'REP_Score', 'OTR_Score', 'NEU_Score']
        for col in score_columns:
            df[col] = 0.0
        
        # Process each text and apply classifier and override rules
        for index, row in df.iterrows():
            text = row['Text']
            # Run classifier
            prediction = classifier(text, candidate_labels, truncation=True, max_length=1024)
            label_scores = {label: score for label, score in zip(prediction['labels'], prediction['scores'])}
            def apply_rule_based_override(text):
                positive_words = ['great', 'well', 'excellent', 'good', 'proud', 'amazing']
                negative_words = ['bad', 'stop', 'disrespectful', 'quiet', 'get out']
                text_lower = text.lower()
                
                if any(word in text_lower for word in positive_words):
                    return 'PRS'
                elif any(word in text_lower for word in negative_words):
                    return 'REP'
                elif text.strip().endswith('?'):
                    return 'OTR'
                return None
            # Apply rule-based override
            override_label = apply_rule_based_override(text)
            if override_label:
                label_scores[override_label] = max(label_scores[override_label], 0.5)  # Override score if higher
            # Update the DataFrame with scores
            for label in candidate_labels:
                df.at[index, f'{label}_Score'] = label_scores[label]
            
        # Determine the label with the highest score
        df['Label'] = df[score_columns].idxmax(axis=1).str.replace('_Score', '')
        new_column_order = [col for col in df.columns if col not in score_columns and col != 'Label'] + score_columns + ['Label']
        df = df[new_column_order]

        def color_map(val):
            """
            Takes a scalar and returns a string with
            the css property `'background-color'` for a color.
            Uses a non-linear scale for color mapping.
            """
            if np.isnan(val):
                return ''
            elif val < 0.2:
                return 'background-color: #ffffcc'  # light yellow
            elif val < 0.3:
                return 'background-color: #d9f0a3'  # light green
            elif val < 0.4:
                return 'background-color: #addd8e'  # green
            elif val < 0.5:
                return 'background-color: #78c679'  # darker green
            else:
                return 'background-color: #31a354'  # dark green
        # Apply the styling
        score_columns = ['PRS_Score', 'REP_Score', 'OTR_Score', 'NEU_Score']
        styled_df = df.style.applymap(color_map, subset=score_columns)

        st.write(styled_df)

        # Create a link for downloading the processed DataFrame
        st.markdown(get_table_download_link(df, score_columns), unsafe_allow_html=True)


def show_know_your_dataset():
    st.title("Know Your Dataset")
    st.write("""
        This section provides insights into the dataset used by ReTeach. 
        Explore data statistics, distributions, and more to understand 
        the foundation of our predictive models.
    """)

def show_labeling_loop():
    st.title("Labeling Loop")
    st.write("""
        In the Labeling Loop, you can actively participate in labeling 
        data points. Your inputs help in improving the accuracy of 
        our machine learning models.
    """)

def main():
    st.sidebar.title("Navigation")
    if st.sidebar.button("Overview"):
        show_overview()
    if st.sidebar.button("AI-Assisted Labeling"):
        show_ai_assisted_labeling()
    if st.sidebar.button("Know your Dataset"):
        show_know_your_dataset()
    if st.sidebar.button("Labeling Loop"):
        show_labeling_loop()


if __name__ == "__main__":
    main()

