import streamlit as st
import numpy as np
import pandas as pd
from io import BytesIO
from transformers import pipeline
import base64
from openpyxl import Workbook
from openpyxl.styles import PatternFill  # Required for Pandas to_excel

def color_map(val):
    """
    Takes a scalar and returns a string with
    the css property `'background-color'` for a color.
    Uses a non-linear scale for color mapping.
    """
    if pd.isna(val):
        return None
    elif val < 0.2:
        return 'FFFFCC'  # light yellow
    elif val < 0.3:
        return 'D9F0A3'  # light green
    elif val < 0.4:
        return 'ADDD8E'  # green
    elif val < 0.5:
        return '78C679'  # darker green
    else:
        return '31A354'  # dark green
    
# Function to convert DataFrame to Excel and then encode it for download
def apply_excel_color_styles(df, color_map_function, target_columns):
    output = BytesIO()
    writer = pd.ExcelWriter(output, engine='openpyxl')
    df.to_excel(writer, index=False, sheet_name='Sheet1')
    workbook = writer.book
    worksheet = writer.sheets['Sheet1']

    for idx, row in df.iterrows():
        for col_idx, col_name in enumerate(df.columns):
            if col_name in target_columns:
                cell = worksheet.cell(row=idx + 2, column=col_idx + 1)
                color = color_map_function(row[col_name])
                if color:
                    cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

    writer.save()
    writer.close()  # Ensure the writer is properly closed
    return output.getvalue()

def get_table_download_link(df, target_columns):
    processed_data = apply_excel_color_styles(df, color_map, target_columns)
    
    b64 = base64.b64encode(processed_data).decode()
    href = f'<a href="data:application/octet-stream;base64,{b64}" download="processed_data.xlsx">Download processed data as Excel</a>'
    return href


# Main function for the Streamlit app
def main():
    st.title("AI-Assisted Labeling :pencil2:")
    st.write("""
        Welcome to the AI-Assisted Labeling page!

        This tool is designed to enhance your data labeling workflow. Follow these simple steps:
        
        1. Upload Your Data: Use the upload button below to upload a CSV file containing your classroom transcripts. 
        2. Review the Labels: Once uploaded, our AI model will automatically suggest labels for your data. These labels are highlighted for easy identification.
        3. Download Labeled Data: After reviewing the AI-suggested labels, you can download the labeled file. 
        
        :clock1: Please wait while we process your file. This can take 30 seconds to 3 minutes based on the size of your file.
    """)

    uploaded_file = st.file_uploader("Choose a file")

    if uploaded_file is not None:
        df = pd.read_csv(uploaded_file)
        if 'Legal' in df.columns:
            df.drop('Legal', axis=1, inplace=True)
        
    
        # Initialize the zero-shot classifier
        classifier = pipeline("zero-shot-classification", model="sileod/deberta-v3-base-tasksource-nli")
        
        # Define candidate labels
        candidate_labels = ['PRS', 'REP', 'OTR', 'NEU']

        # Initialize columns for scores
        score_columns = ['PRS_Score', 'REP_Score', 'OTR_Score', 'NEU_Score']
        for col in score_columns:
            df[col] = 0.0
        
        # Process each text and apply classifier and override rules
        for index, row in df.iterrows():
            text = row['Text']
            # Run classifier
            prediction = classifier(text, candidate_labels, truncation=True, max_length=1024)
            label_scores = {label: score for label, score in zip(prediction['labels'], prediction['scores'])}
            def apply_rule_based_override(text):
                positive_words = ['great', 'well', 'excellent', 'good', 'proud', 'amazing']
                negative_words = ['bad', 'stop', 'disrespectful', 'quiet', 'get out']
                text_lower = text.lower()
                
                if any(word in text_lower for word in positive_words):
                    return 'PRS'
                elif any(word in text_lower for word in negative_words):
                    return 'REP'
                elif text.strip().endswith('?'):
                    return 'OTR'
                return None
            # Apply rule-based override
            override_label = apply_rule_based_override(text)
            if override_label:
                label_scores[override_label] = max(label_scores[override_label], 0.5)  # Override score if higher
            # Update the DataFrame with scores
            for label in candidate_labels:
                df.at[index, f'{label}_Score'] = label_scores[label]
            
        # Determine the label with the highest score
        df['Label'] = df[score_columns].idxmax(axis=1).str.replace('_Score', '')
        new_column_order = [col for col in df.columns if col not in score_columns and col != 'Label'] + score_columns + ['Label']
        df = df[new_column_order]

        def color_map(val):
            """
            Takes a scalar and returns a string with
            the css property `'background-color'` for a color.
            Uses a non-linear scale for color mapping.
            """
            if np.isnan(val):
                return ''
            elif val < 0.2:
                return 'background-color: #ffffcc'  # light yellow
            elif val < 0.3:
                return 'background-color: #d9f0a3'  # light green
            elif val < 0.4:
                return 'background-color: #addd8e'  # green
            elif val < 0.5:
                return 'background-color: #78c679'  # darker green
            else:
                return 'background-color: #31a354'  # dark green
        # Apply the styling
        score_columns = ['PRS_Score', 'REP_Score', 'OTR_Score', 'NEU_Score']
        styled_df = df.style.applymap(color_map, subset=score_columns)

        st.write(styled_df)

        # Create a link for downloading the processed DataFrame
        st.markdown(get_table_download_link(df, score_columns), unsafe_allow_html=True)

if __name__ == "__main__":
    main()

